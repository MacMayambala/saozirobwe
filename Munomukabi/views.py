from django.shortcuts import render

# views.py
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from .forms import MemberForm  # If using ModelForm
# from .forms import MemberRegistrationForm  # If using a regular Form
from .models import Member
from django.shortcuts import render, redirect
#from .forms import MemberRegistrationForm
from .models import Parish, Village
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import MemberRegistrationForm
from .models import Member
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import MemberForm
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import MemberForm  # Ensure you have imported your form
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import Member, ServedMember
from .forms import ServeMemberForm
from django.contrib import messages


from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import Member, ServedMember
from .forms import ServeMemberForm
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import Member
from datetime import date
from .models import Member

from django.http import JsonResponse
from django.shortcuts import render
from datetime import date
from .models import Member
# In your views.py file
from django.http import JsonResponse
from .models import Member





from django.urls import reverse
from django.shortcuts import render, get_object_or_404, redirect
from .models import Member
from .forms import MemberForm  # Create a Django form

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponseRedirect
from .models import Customer
from .forms import MemberForm

# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Customer
@login_required
def memberSum(request, member_id):
    # Ensure the ID exists
    member = get_object_or_404(Customer, cus_id=member_id)

    if request.method == "POST":
        # Update member with new images if provided
        if 'profile_picture' in request.FILES:
            member.profile_picture = request.FILES['profile_picture']
        if 'signature_photo' in request.FILES:
            member.signature_photo = request.FILES['signature_photo']
        if 'id_scan' in request.FILES:
            member.id_scan = request.FILES['id_scan']
        member.save()
        messages.success(request, "Images updated successfully!")
        return redirect("memberSum", member_id=member.cus_id)  # Redirect to the same page after saving

    return render(request, "Member_Sum.html", {"member": member})





from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import date, timedelta
from .models import Customer, Member
from .forms import RenewalForm


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import Customer, Member
from .forms import RenewalForm
from datetime import date, timedelta
from django.db import IntegrityError
# Munomukabi/views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.db import IntegrityError
from .models import Customer, Member
from .forms import MunoMemberForm, RenewalForm
from datetime import date, timedelta

@login_required
def muno_member_detail(request, cus_id):
    customer = get_object_or_404(Customer, cus_id=cus_id)
    member = Member.objects.filter(customer=customer).first()
    if not member:
        messages.error(request, "No member record found for this customer.")
        return redirect('mono_list')
    
    form = RenewalForm(initial={'subscription_start': date.today()})
    return render(request, 'munomukabitemp/muno_details.html', {
        'form': form,
        'member': member,
        'customer': customer,
    })

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Member
from datetime import date, timedelta
from django.db import IntegrityError
from .forms import RenewalForm  # Assuming RenewalForm is defined elsewhere

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Member
from datetime import date, timedelta
from django.db import transaction, IntegrityError
from .forms import RenewalForm

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Member
from datetime import date, timedelta
from django.db import transaction, IntegrityError
from .forms import RenewalForm

@login_required
def renew_sub(request, member_id):
    member = get_object_or_404(Member, id=member_id)

    # Check if member is expired
    if member.status != 'expired':
        messages.error(request, "This member's subscription is still active and cannot be renewed.")
        return redirect("mono_list")  # Redirect to /muno/ without member_id

    # Check renewal permission
    if not member.can_renew(request.user):
        messages.error(request, "Renewal is allowed only once per year unless authorized by an admin.")
        return redirect("mono_list")  # Redirect to /muno/ without member_id

    if request.method == "POST":
        # Handle confirmation submission
        if request.POST.get('confirm_submission'):
            messages.success(request, f"Subscription for {member.customer.first_name} {member.customer.surname} renewed successfully.")
            return redirect("mono_list")  # Redirect to /muno/

        form = RenewalForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    subscription_start = form.cleaned_data['subscription_start']
                    payment_date = form.cleaned_data['payment_date']
                    payment_number = form.cleaned_data['payment_number']
                    booklet_number = form.cleaned_data['booklet_number']
                    teller = form.cleaned_data['teller']

                    # Create new member record
                    new_member = Member(
                        customer=member.customer,
                        subscription_start=subscription_start,
                        subscription_end=subscription_start + timedelta(days=365),
                        status='active',
                        renewal_count=member.renewal_count + 1,
                        last_renewal_date=date.today(),
                        payment_date=payment_date,
                        payment_number=payment_number,
                        booklet_number=booklet_number,
                        teller=teller,
                        spouse_name=member.spouse_name,
                        spouse_status=member.spouse_status,
                        spouse_address=member.spouse_address,
                        mother_name=member.mother_name,
                        mother_village=member.mother_village,
                        mother_status=member.mother_status,
                        mother_guardian=member.mother_guardian,
                        mother_guardian_address=member.mother_guardian_address,
                        father_name=member.father_name,
                        father_village=member.father_village,
                        father_status=member.father_status,
                        father_guardian=member.father_guardian,
                        father_guardian_address=member.father_guardian_address,
                        child1_name=member.child1_name,
                        child2_name=member.child2_name,
                        next_of_kin_Name=member.next_of_kin_Name,
                        next_of_kin_relationship=member.next_of_kin_relationship,
                        next_of_kin_village=member.next_of_kin_village,
                        next_of_kin_phone=member.next_of_kin_phone,
                        sao_officer=member.sao_officer,
                    )
                    new_member.save()

                    # Update old member's status
                    member.status = 'archived'
                    member.save()

                # Show confirmation modal
                return render(request, 'munomukabitemp/muno_details.html', {
                    'member': member,
                    'form': form,
                    'show_confirmation': True,
                    'submitted_data': {
                        'subscription_start': subscription_start,
                        'payment_date': payment_date,
                        'payment_number': payment_number,
                        'booklet_number': booklet_number,
                        'teller': teller,
                    }
                })

            except IntegrityError as e:
                error_message = str(e).lower()
                if 'unique_active_member_per_customer' in error_message:
                    messages.error(request, "An active member already exists for this customer.")
                elif 'payment_number' in error_message:
                    form.add_error('payment_number', "This payment number is already in use.")
                elif 'booklet_number' in error_message:
                    form.add_error('booklet_number', "This booklet number is already in use.")
                else:
                    messages.error(request, "An unexpected error occurred while renewing the subscription.")
                return render(request, 'munomukabitemp/muno_details.html', {
                    'form': form,
                    'member': member,
                    'show_modal': True
                })
        else:
            messages.error(request, "Please correct the errors in the form below.")
            return render(request, 'munomukabitemp/muno_details.html', {
                'form': form,
                'member': member,
                'show_modal': True
            })
    else:
        form = RenewalForm(initial={'subscription_start': date.today()}, user=request.user)

    return render(request, 'munomukabitemp/muno_details.html', {
        'form': form,
        'member': member,
        'show_modal': False  # No modal on initial GET
    })
@login_required
def register_muno(request, cus_id):
    customer = get_object_or_404(Customer, cus_id=cus_id)
    member = Member.objects.filter(customer=customer, status='active').first()

    if request.method == "POST":
        form = MunoMemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            try:
                member_instance = form.save(commit=False)
                member_instance.customer = customer
                if not member:  # New registration
                    member_instance.status = 'active'
                    member_instance.renewal_count = 1
                    if member_instance.subscription_start:
                        member_instance.last_renewal_date = member_instance.subscription_start
                    Member.objects.filter(customer=customer).exclude(status='active').update(status='archived')
                member_instance.save()
                action = "updated" if member else "registered"
                messages.success(request, f"Member details {action} successfully.")
                return redirect("member_detail", cus_id=customer.cus_id)
            except IntegrityError as e:
                error_message = str(e).lower()
                if 'unique_active_member_per_customer' in error_message:
                    messages.error(request, "An active member already exists for this customer.")
                elif 'payment_number' in error_message:
                    messages.error(request, "The payment number is already in use.")
                elif 'booklet_number' in error_message:
                    messages.error(request, "The booklet number is already in use.")
                else:
                    messages.error(request, "An error occurred while saving the member details.")
                print("IntegrityError:", error_message)
        else:
            messages.error(request, "Please correct the errors below.")
            print("Form errors:", form.errors.as_data())
    else:
        form = MunoMemberForm(instance=member)

    return render(request, 'Munomukabi/registermunomukabi.html', {
        'form': form,
        'member': member,
        'customer': customer,
    })
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import ServedMember

# Munomukabi/views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import ServedMember
# Munomukabi/views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import ServedMember

from django.contrib import messages
# Munomukabi/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Member, Customer, ServedMember  # Adjust to your models
from django.shortcuts import get_object_or_404

@login_required
def serve_member(request, member_id):
    # Try to fetch the member based on the URL parameter
    try:
        member = Member.objects.get(id=member_id)
        customer = member.customer
    except Member.DoesNotExist:
        messages.error(request, "Member not found.")
        return redirect('served_members_report')

    # Handle POST request (service registration)
    if request.method == 'POST':
        amount_given = request.POST.get('amount_given')
        reason = request.POST.get('reason')

        try:
            served_member = ServedMember(
                member=member,
                amount_given=amount_given,
                reason=reason,
                served_by=request.user
            )
            served_member.save()
            messages.success(request, f"Member {customer.first_name} {customer.surname} Served.")
            return redirect('served_members_report')
        except ValueError as e:
            return render(request, 'serve_member.html', {
                'customer': customer,
                'errors': {'amount_given': str(e) if 'amount_given' in str(e) else 'Invalid input'}
            })

    # Render the form for GET request
    return render(request, 'serve_member.html', {'customer': customer})

# munomukabitemp/views.py
# munomukabitemp/views.py
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import ServedMember

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
import csv
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from io import BytesIO

@login_required
def served_members_report(request):
    # Get query parameters
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    sort_field = request.GET.get('sort_field', 'date_served')
    sort_order = request.GET.get('sort_order', 'desc')

    # Validate sort field
    valid_sort_fields = ['name', 'member_number', 'amount_given', 'reason', 'date_served', 'served_by']
    if sort_field not in valid_sort_fields:
        sort_field = 'date_served'

    # Base queryset
    served_members = ServedMember.objects.select_related('member__customer', 'served_by')

    # Apply search filter
    if search_query:
        served_members = served_members.filter(
            Q(member__customer__first_name__icontains=search_query) |
            Q(member__customer__surname__icontains=search_query) |
            Q(member__customer__member_number__icontains=search_query) |
            Q(reason__icontains=search_query)
        )

    # Apply date range filter
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            served_members = served_members.filter(date_served__gte=start_date)
        except ValueError:
            pass
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            served_members = served_members.filter(date_served__lte=end_date)
        except ValueError:
            pass

    # Apply sorting
    order_prefix = '-' if sort_order == 'desc' else ''
    if sort_field == 'name':
        # Sort by first_name and surname
        served_members = served_members.order_by(
            f'{order_prefix}member__customer__first_name',
            f'{order_prefix}member__customer__surname'
        )
    elif sort_field == 'member_number':
        served_members = served_members.order_by(f'{order_prefix}member__customer__member_number')
    elif sort_field == 'served_by':
        served_members = served_members.order_by(f'{order_prefix}served_by__username')
    else:
        served_members = served_members.order_by(f'{order_prefix}{sort_field}')

    # Calculate summary data
    total_amount = served_members.aggregate(Sum('amount_given'))['amount_given__sum'] or 0
    date_range = (
        f"{start_date.strftime('%Y-%m-%d') if start_date else 'Start'} - {end_date.strftime('%Y-%m-%d') if end_date else 'End'}"
        if start_date or end_date else 'All Time'
    )

    # Handle exports
    if request.path.endswith('/export-csv/'):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="served_members_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Member Number', 'Amount Given (UGX)', 'Reason', 'Date Served', 'Served By'])
        for service in served_members:
            writer.writerow([
                f"{service.member.customer.first_name} {service.member.customer.surname}",
                service.member.customer.member_number or 'N/A',
                service.amount_given,
                service.reason,
                service.date_served.strftime('%Y-%m-%d'),
                service.served_by.username if service.served_by else 'Unknown'
            ])
        return response
    elif request.path.endswith('/export-pdf/'):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title="Served Members Report")
        data = [['Name', 'Member Number', 'Amount Given (UGX)', 'Reason', 'Date Served', 'Served By']]
        for service in served_members:
            data.append([
                f"{service.member.customer.first_name} {service.member.customer.surname}",
                service.member.customer.member_number or 'N/A',
                str(service.amount_given),
                service.reason,
                service.date_served.strftime('%Y-%m-%d'),
                service.served_by.username if service.served_by else 'Unknown'
            ])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e4d92')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        doc.build([table])
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="served_members_report.pdf"'
        buffer.close()
        return response

    # Pagination
    paginator = Paginator(served_members, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            'served_members': [
                {
                    'member': {
                        'customer': {
                            'first_name': service.member.customer.first_name,
                            'surname': service.member.customer.surname,
                            'member_number': service.member.customer.member_number or 'N/A',
                        }
                    },
                    'amount_given': str(service.amount_given),
                    'reason': service.reason,
                    'date_served': service.date_served.strftime('%Y-%m-%d'),
                    'served_by': {'username': service.served_by.username if service.served_by else 'Unknown'},
                }
                for service in page_obj
            ],
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'total_amount': str(total_amount),
            'date_range': date_range
        }
        return JsonResponse(data)

    return render(request, 'served_report.html', {'served_members': page_obj})

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Customer

@login_required
def search_member(request):
    search_query = request.GET.get('search', '').strip()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Filter by member number, first name, or surname
            members = Customer.objects.filter(
                Q(member_number__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(surname__icontains=search_query)
            )

            # Pagination (limit results to 5 per page)
            paginator = Paginator(members, 5)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            members_data = [{
                'cus_id': member.cus_id,
                'first_name': member.first_name,
                'surname': member.surname,
                'member_number': member.member_number,
                'branch': member.branch,
                'gender': member.gender,
                'village': member.village
            } for member in page_obj]

            return JsonResponse({
                'success': True,
                'members': members_data,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

    

@login_required
def update_member(request):
    if request.method == "POST":
        form = MemberForm(request.POST, request.FILES)
        print(request.POST)  # Debugging: Check what data is being submitted

        if form.is_valid():
            form.save()
            messages.success(request, "Member details updated successfully!")
            return redirect('register')  # Redirect after successful submission
        else:
            print("Form Errors:", form.errors)  # Print errors to console
            messages.error(request, "Error in form submission. Please check the fields.")

    else:
        form = MemberForm()

    return render(request, "register1.html", {"form": form})  

@login_required
def register_member(request):
    if request.method == "POST":
        form = MemberForm(request.POST, request.FILES)
        print(request.POST)  # Debugging: Check what data is being submitted

        if form.is_valid():
            form.save()
            messages.success(request, "Member registered successfully!")
            return redirect('register')  # Redirect after successful submission
        else:
            print("Form Errors:", form.errors)  # Print errors to console
            messages.error(request, "Error in form submission. Please check the fields.")

    else:
        form = MemberForm()

    return render(request, "register1.html", {"form": form})




# AJAX endpoint for dynamic dropdowns
from django.http import JsonResponse
@login_required
def load_parishes(request):
    zone_id = request.GET.get('zone_id')
    parishes = Parish.objects.filter(zone_id=zone_id).values('id', 'name')
    return JsonResponse(list(parishes), safe=False)
@login_required
def load_villages(request):
    parish_id = request.GET.get('parish_id')
    villages = Village.objects.filter(parish_id=parish_id).values('id', 'name')
    return JsonResponse(list(villages), safe=False)



@login_required
def reg(request):
    return render(request,"register.html")



from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Member
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import io
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Member
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import io

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, F
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from .models import Member
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import io
from datetime import datetime

@login_required
def members_list(request):
    # Get query parameters
    name = request.GET.get('name', '')
    member_number = request.GET.get('member_number', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    sort_field = request.GET.get('sort_field', 'subscription_start')
    sort_order = request.GET.get('sort_order', 'desc')

    # Validate sort field
    valid_sort_fields = ['name', 'member_number', 'phone', 'dob', 'subscription_start', 'subscription_end', 'branch', 'status']
    if sort_field not in valid_sort_fields:
        sort_field = 'subscription_start'

    # Base queryset
    members = Member.objects.select_related('customer')

    # Apply filters
    if name:
        members = members.filter(
            Q(customer__first_name__icontains=name) |
            Q(customer__surname__icontains=name)
        )
    if member_number:
        members = members.filter(customer__member_number__icontains=member_number)
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            members = members.filter(subscription_start__gte=start_date)
        except ValueError:
            pass
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            members = members.filter(subscription_end__lte=end_date)
        except ValueError:
            pass

    # Apply sorting
    order_prefix = '-' if sort_order == 'desc' else ''
    if sort_field == 'name':
        members = members.order_by(
            f'{order_prefix}customer__first_name',
            f'{order_prefix}customer__surname'
        )
    elif sort_field == 'member_number':
        members = members.order_by(f'{order_prefix}customer__member_number')
    elif sort_field == 'phone':
        members = members.order_by(f'{order_prefix}customer__phone')  # Updated field
    elif sort_field == 'dob':
        members = members.order_by(f'{order_prefix}customer__dob')
    elif sort_field == 'branch':
        members = members.order_by(f'{order_prefix}customer__branch__name')  # Adjust if 'name' isn't correct
    else:
        members = members.order_by(f'{order_prefix}{sort_field}')

    # Calculate summary data
    total_items = members.count()
    date_range = (
        f"{start_date.strftime('%Y-%m-%d') if start_date else 'Start'} - {end_date.strftime('%Y-%m-%d') if end_date else 'End'}"
        if start_date or end_date else 'All Time'
    )

    # Handle AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        paginator = Paginator(members, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        data = {
            'members': [
                {
                    'customer': {
                        'first_name': member.customer.first_name,
                        'surname': member.customer.surname,
                        'member_number': member.customer.member_number or 'N/A',
                        'phone': member.customer.phone or 'N/A',  # Updated field
                        'dob': member.customer.dob.strftime('%Y-%m-%d') if member.customer.dob else 'N/A',
                        'branch': str(member.customer.branch) if member.customer.branch else 'N/A',
                    },
                    'subscription_start': member.subscription_start.strftime('%Y-%m-%d') if member.subscription_start else 'N/A',
                    'subscription_end': member.subscription_end.strftime('%Y-%m-%d') if member.subscription_end else 'N/A',
                    'status': member.status or 'N/A',
                }
                for member in page_obj
            ],
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': total_items,
            'date_range': date_range
        }
        return JsonResponse(data)

    # Handle export requests
    if request.path.endswith('/export-members-to-excel/'):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Members Report'
        headers = ['Name', 'Member Number', 'Phone', 'Date of Birth', 'Subscription Start', 'Subscription End', 'Branch', 'Status']
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for member in members:
            worksheet.append([
                f"{member.customer.first_name} {member.customer.surname}",
                member.customer.member_number or 'N/A',
                member.customer.phone or 'N/A',  # Updated field
                member.customer.dob.strftime('%Y-%m-%d') if member.customer.dob else 'N/A',
                member.subscription_start.strftime('%Y-%m-%d') if member.subscription_start else 'N/A',
                member.subscription_end.strftime('%Y-%m-%d') if member.subscription_end else 'N/A',
                str(member.customer.branch) if member.customer.branch else 'N/A',
                member.status or 'N/A'
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="members_report.xlsx"'
        workbook.save(response)
        return response
    elif request.path.endswith('/export-members-to-pdf/'):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, title="Members Report")
        data = [['Name', 'Member Number', 'Phone', 'Date of Birth', 'Subscription Start', 'Subscription End', 'Branch', 'Status']]
        for member in members:
            data.append([
                f"{member.customer.first_name} {member.customer.surname}",
                member.customer.member_number or 'N/A',
                member.customer.phone or 'N/A',  # Updated field
                member.customer.dob.strftime('%Y-%m-%d') if member.customer.dob else 'N/A',
                member.subscription_start.strftime('%Y-%m-%d') if member.subscription_start else 'N/A',
                member.subscription_end.strftime('%Y-%m-%d') if member.subscription_end else 'N/A',
                str(member.customer.branch) if member.customer.branch else 'N/A',
                member.status or 'N/A'
            ])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e4d92')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        doc.build([table])
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="members_report.pdf"'
        buffer.close()
        return response

    # Render template
    paginator = Paginator(members, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    context = {
        'members': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'reports/all_members.html', context)
@require_GET
def filter_members(request):
    members = Member.objects.all().order_by("id")
  # ✅ Ordering added
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    name = request.GET.get("name")
    member_number = request.GET.get("member_number")

    if start_date:
        members = members.filter(subscription_start__gte=start_date)
    if end_date:
        members = members.filter(subscription_end__lte=end_date)
    if name:
        members = members.filter(customer__surname__icontains=name)
    if member_number:
        members = members.filter(customer__member_number__icontains=member_number)

    paginator = Paginator(members, 5)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    data = {
        "members": [{
            "id": m.id,
            "customer": {
                "surname": m.customer.surname,
                "member_number": m.customer.member_number,
                "phone": m.customer.phone or "",
                "dob": m.customer.dob.strftime("%Y-%m-%d") if m.customer.dob else "",
                "branch": m.customer.branch
            },
            "subscription_start": m.subscription_start.strftime("%Y-%m-%d") if m.subscription_start else "",
            "subscription_end": m.subscription_end.strftime("%Y-%m-%d") if m.subscription_end else "",
            "status": m.status
        } for m in page_obj],
        "has_previous": page_obj.has_previous(),
        "has_next": page_obj.has_next(),
        "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
        "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
        "current_page": page_obj.number,
        "total_pages": paginator.num_pages
    }
    return JsonResponse(data)

@login_required
def export_members_to_excel(request):
    members = Member.objects.all()
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    name = request.GET.get("name")
    member_number = request.GET.get("member_number")

    if start_date:
        members = members.filter(subscription_start__gte=start_date)
    if end_date:
        members = members.filter(subscription_end__lte=end_date)
    if name:
        members = members.filter(customer__surname__icontains=name)
    if member_number:
        members = members.filter(customer__member_number__icontains=member_number)

    # No pagination here - we want all filtered members
    members = members  # Simply use the filtered queryset directly

    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    # Headers
    headers = ["ID", "Name", "Account Number", "Phone", "Date of Birth", "Subscription Start", "Subscription End", "Branch", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data - Iterate over all filtered members
    for row, member in enumerate(members, start=2):
        ws.append([
            member.id,
            member.customer.surname,
            member.customer.member_number,
            member.customer.phone or "",
            member.customer.dob.strftime("%Y-%m-%d") if member.customer.dob else "",
            member.subscription_start.strftime("%Y-%m-%d") if member.subscription_start else "",
            member.subscription_end.strftime("%Y-%m-%d") if member.subscription_end else "",
            member.customer.branch,
            member.status
        ])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Prepare response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=members.xlsx"
    wb.save(response)
    return response

def export_members_to_pdf(request):
    members = Member.objects.all()
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    name = request.GET.get("name")
    member_number = request.GET.get("member_number")

    if start_date:
        members = members.filter(subscription_start__gte=start_date)
    if end_date:
        members = members.filter(subscription_end__lte=end_date)
    if name:
        members = members.filter(customer__surname__icontains=name)
    if member_number:
        members = members.filter(customer__member_number__icontains=member_number)

    # No pagination here - we want all filtered members
    members = members  # Use the full filtered queryset

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Table data
    data = [["ID", "Name", "Account Number", "Phone", "Date of Birth", "Subscription Start", "Subscription End", "Branch", "Status"]]
    for member in members:
        data.append([
            str(member.id),
            member.customer.surname,
            member.customer.member_number,
            member.customer.phone or "",
            member.customer.dob.strftime("%Y-%m-%d") if member.customer.dob else "",
            member.subscription_start.strftime("%Y-%m-%d") if member.subscription_start else "",
            member.subscription_end.strftime("%Y-%m-%d") if member.subscription_end else "",
            member.customer.branch,
            member.status
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    # Prepare response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=members.pdf"
    response.write(pdf)
    return response




from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib import messages
from .models import Customer, Member
from .forms import MemberRegistrationForm
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import io

@login_required
def member_form(request, cus_id=None):
    if request.method == "POST":
        form = MemberRegistrationForm(request.POST)
        if form.is_valid():
            if cus_id:  # Update existing member
                try:
                    member = Member.objects.get(customer__id=cus_id)
                    customer = member.customer
                    # Update Customer fields (excluding read-only branch and member_number)
                    customer.surname = form.cleaned_data['surname']
                    customer.phone = form.cleaned_data['phone']
                    customer.dob = form.cleaned_data['dob']
                    customer.save()

                    # Update Member fields
                    member.subscription_start = form.cleaned_data['subscription_start']
                    member.subscription_end = form.cleaned_data['subscription_end']
                    member.status = form.cleaned_data['status']
                    member.save()

                    messages.success(request, "Member updated successfully!")
                except Member.DoesNotExist:
                    messages.error(request, "Member not found.")
            else:  # Create new member
                customer = Customer.objects.create(
                    surname=form.cleaned_data['surname'],
                    member_number=form.cleaned_data['member_number'],
                    phone=form.cleaned_data['phone'],
                    dob=form.cleaned_data['dob'],
                    branch=form.cleaned_data['branch']
                )
                Member.objects.create(
                    customer=customer,
                    subscription_start=form.cleaned_data['subscription_start'],
                    subscription_end=form.cleaned_data['subscription_end'],
                    status=form.cleaned_data['status']
                )
                messages.success(request, "Member registered successfully!")
            return redirect('member_list')
    else:
        if cus_id:  # Pre-fill form for existing member
            try:
                member = Member.objects.get(customer__id=cus_id)
                form = MemberRegistrationForm(instance=member)
            except Member.DoesNotExist:
                form = MemberRegistrationForm()
                messages.error(request, "Member not found for this Customer ID.")
        else:  # New member form
            form = MemberRegistrationForm()

    context = {
        'form': form,
        'cus_id': cus_id
    }
    return render(request, "member_form.html", context)


#################################################################################################################################
# your_app_name/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Customer, Member
from .forms import MunoMemberForm


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Customer, Member
from .forms import MunoMemberForm

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Customer, Member
from .forms import MunoMemberForm


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Customer, Member
from .forms import MunoMemberForm
from django.db import IntegrityError

@login_required
def register_muno(request, cus_id):
    customer = get_object_or_404(Customer, cus_id=cus_id)
    # Prioritize active Member, fall back to any Member
    member = Member.objects.filter(customer=customer, status='active').first() or Member.objects.filter(customer=customer).first()

    if request.method == "POST":
        # If no active member exists, create a new one; otherwise, update the active one
        instance = member if member and member.status == 'active' else None
        form = MunoMemberForm(request.POST, request.FILES, instance=instance, customer=customer)
        if form.is_valid():
            try:
                member_instance = form.save(commit=False)
                member_instance.customer = customer
                if not instance:  # New record
                    member_instance.status = 'active'
                    # Ensure renewal_count starts at 1 for new records
                    member_instance.renewal_count = 1
                    # Set last_renewal_date for new records
                    if member_instance.subscription_start:
                        member_instance.last_renewal_date = member_instance.subscription_start
                    # If there's an existing non-active member, mark it as archived
                    if member and member.status != 'active':
                        member.status = 'archived'
                        member.save()
                member_instance.save()
                action = "updated" if instance else "registered"
                messages.success(request, f"Member details {action} successfully.")
                return redirect("register_muno", cus_id=customer.cus_id)
            except IntegrityError as e:
                if 'unique_active_member_per_customer' in str(e):
                    messages.error(request, "An active member already exists for this customer.")
                elif 'payment_number' in str(e):
                    messages.error(request, "The payment number is already in use.")
                elif 'booklet_number' in str(e):
                    messages.error(request, "The booklet number is already in use.")
                else:
                    messages.error(request, "An error occurred while saving the member details.")
                print("IntegrityError:", str(e))
        else:
            messages.error(request, "Please correct the errors below.")
            print("Form errors:", form.errors)
    else:
        form = MunoMemberForm(instance=member, customer=customer)

    return render(request, 'registermunomukabi.html', {
        'form': form,
        'member': member,
        'customer': customer,
    })



@login_required
def member_register(request):
    """Register a new member (standalone, not tied to a specific cus_id)"""
    if request.method == "POST":
        form = MunoMemberForm(request.POST, request.FILES)
        if form.is_valid():
            member = form.save()
            messages.success(request, "Member registered successfully.")
            return redirect("member_detail", cus_id=member.customer.cus_id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = MunoMemberForm()

    return render(request, 'registermunomukabi.html', {
        'form': form,
    })


#################################################################################
# Munomukabi/views.py
from django.shortcuts import render, redirect
from django.utils import timezone
from Munomukabi.models import CronJobStatus
from Munomukabi.management.commands.expiry_notification import Command as NotifyCommand
import threading
from django.http import JsonResponse

def cron_status_view(request):
    # Get or create status object
    cron_status, created = CronJobStatus.objects.get_or_create(id=1)
    return render(request, 'cronjob/cron_status.html', {'cron_status': cron_status})

def start_cron_job(request):
    cron_status, created = CronJobStatus.objects.get_or_create(id=1)

    if not cron_status.running:
        def run_task():
            cron_status.start()
            NotifyCommand().handle()  # This runs your notify_expired_members code!
            cron_status.finish()

        threading.Thread(target=run_task).start()

    return JsonResponse({'status': 'started'})

def cron_status_api(request):
    cron_status, created = CronJobStatus.objects.get_or_create(id=1)
    return JsonResponse({
        'running': cron_status.running,
        'last_run': cron_status.last_run.strftime("%d %b %Y, %H:%M:%S") if cron_status.last_run else "Never"
    })
#####################################################################################################################
# Munomukabi list
from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Member  # Adjust to your actual model


from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Member  # Adjust import based on your app structure
import logging
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Member

logger = logging.getLogger(__name__)

import logging
from django.shortcuts import render
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.db.models import Q
from .models import Member

logger = logging.getLogger(__name__)
@login_required
def munolist(request):
    search_query = request.GET.get('search', '').strip()
    logger.debug(f"Search query: {search_query}, Page: {request.GET.get('page', 1)}")

    # Fetch members with search filtering
    members = Member.objects.all()
    if search_query:
        members = members.filter(
            Q(customer__first_name__icontains=search_query) |
            Q(customer__surname__icontains=search_query) |
            Q(customer__member_number__icontains=search_query)
        )
    
    # Paginate results
    paginator = Paginator(members, 5)  # 5 members per page
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except Exception as e:
        logger.error(f"Pagination error: {str(e)}", exc_info=True)
        page_obj = paginator.page(1)  # Fallback to page 1

    # Handle AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            members_data = [
                {
                    'id': member.id,
                    'customer': {
                        'first_name': member.customer.first_name,
                        'surname': member.customer.surname,
                        'member_number': member.customer.member_number,
                        'phone': member.customer.phone or '',
                        'branch': str(member.customer.branch) if member.customer.branch else '',
                        'cus_id': member.customer.cus_id,  # Include cus_id for detail URL
                    },
                    'subscription_end': member.subscription_end.strftime('%Y-%m-%d') if member.subscription_end else '',
                    'status': member.status,
                } for member in page_obj
            ]
            return JsonResponse({
                'members': members_data,
                'has_previous': page_obj.has_previous(),
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
                'has_next': page_obj.has_next(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages
            })
        except Exception as e:
            logger.error(f"Error serializing members data: {str(e)}", exc_info=True)
            return JsonResponse({
                'error': 'Failed to load data due to serialization error.',
            }, status=500)
    
    context = {
        'members': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'munomukabitemp/muno_list.html', context)

@login_required
def filter_members(request):
    search_query = request.GET.get('search', '').strip()
    members = Member.objects.all()
    
    if search_query:
        members = members.filter(
            Q(customer__first_name__icontains=search_query) |
            Q(customer__surname__icontains=search_query) |
            Q(customer__member_number__icontains=search_query)
        )
    
    paginator = Paginator(members, 5)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    members_data = [
        {
            'id': member.id,  # Required for View button URL
            'customer': {
                'first_name': member.customer.first_name,
                'surname': member.customer.surname,
                'member_number': member.customer.member_number,
                'phone': member.customer.phone,
                'dob': member.customer.dob.strftime('%Y-%m-%d') if member.customer.dob else '',
                'branch': member.customer.branch,
            },
            'subscription_start': member.subscription_start.strftime('%Y-%m-%d') if member.subscription_start else '',
            'subscription_end': member.subscription_end.strftime('%Y-%m-%d') if member.subscription_end else '',
            'status': member.status,
        }
        for member in page_obj
    ]
    
    data = {
        'members': members_data,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages,
    }
    
    return JsonResponse(data)

@login_required
def view_member_details(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    
    return render(request, 'munomukabitemp/muno_details.html', {'member': member})



#######################################################################
import logging
from django.shortcuts import render
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Customer  # Adjust import based on your app structure

logger = logging.getLogger(__name__)

def member_list(request):
    try:
        customers = Customer.objects.all().order_by('cus_id')
        logger.debug(f"Loaded {customers.count()} customers for initial render")
        context = {'customers': customers}
        return render(request, 'members.html', context)  # Adjust template path
    except Exception as e:
        logger.error(f"Error in members view: {str(e)}", exc_info=True)
        return render(request, 'members.html', {'error': 'An error occurred while loading customers.'}, status=500)

def search_members(request):
    try:
        query = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        try:
            page = int(page)
        except ValueError:
            page = 1

        logger.debug(f"Search query: {query}, Page: {page}")

        customers = Customer.objects.all()
        if query:
            customers = customers.filter(
                Q(first_name__icontains=query) |
                Q(surname__icontains=query) |
                Q(member_number__icontains=query) |
                Q(cus_id__icontains=query) |
                Q(phone__icontains=query)
            )

        customers = customers.order_by('cus_id')
        logger.debug(f"Found {customers.count()} customers")

        paginator = Paginator(customers, 5)
        page_obj = paginator.get_page(page)

        data = {
            'customers': [
                {
                    'cus_id': customer.cus_id,
                    'first_name': customer.first_name,
                    'surname': customer.surname,
                    'member_number': customer.member_number,
                    'gender': customer.gender,
                    'age': customer.age,
                    'village': customer.village,
                    'phone': customer.phone,
                    'profile_picture': customer.profile_picture.url if customer.profile_picture else None,
                }
                for customer in page_obj
            ],
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }

        logger.debug(f"Returning {len(data['customers'])} customers in response")
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error in search_members view: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Internal server error'}, status=500)