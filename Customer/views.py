from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import CustomerForm
@login_required
def registermember(request):
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():  # Checks if the form passes validation
            form.save()
            messages.success(request, "Member registered successfully!")
            return redirect('member_registration')  # Redirect to a success page
        else:
            return render(request, 'register.html', {'form': form})  # Re-render with errors
    else:
        form = CustomerForm()
    return render(request, 'register.html', {'form': form})


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Customer
from django.contrib import messages
@login_required
def member_summary(request, cus_id):
    member = get_object_or_404(Customer, cus_id=cus_id)

    if request.method == 'POST':
        # Update member with new images if provided
        if 'profile_picture' in request.FILES:
            member.profile_picture = request.FILES['profile_picture']
        if 'signature_photo' in request.FILES:
            member.signature_photo = request.FILES['signature_photo']
        if 'id_scan' in request.FILES:
            member.id_scan = request.FILES['id_scan']
        member.save()
        messages.success(request, "Images updated successfully!")
        return redirect('member_summary', cus_id=member.cus_id)

    return render(request, 'Member_Sum.html', {'member': member})




# yourapp/views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Customer
from .forms import CustomerEditForm
from django.contrib import messages
@login_required
def edit_customer(request, cus_id):
    customer = get_object_or_404(Customer, cus_id=cus_id)
    
    if request.method == 'POST':
        form = CustomerEditForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f"Member {customer.first_name}   {customer.surname} updated successfully!")
            return redirect('members')  # Redirect to your members list or another page
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CustomerEditForm(instance=customer)
    
    return render(request, 'edit_customer.html', {
        'form': form,
        'customer': customer,
    })



##########################################################################
@login_required
def download_template(request):
    """Generate and download a sample Excel template for importing customers"""
    import pandas as pd
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font

    # Required fields list
    required_fields = [
        'branch', 'mem_reg_date', 'member_number', 'first_name',
        'surname', 'gender', 'dob', 'id_number', 'card_number',
        'phone', 'kin_sname', 'kin_fname', 'kin_phone'
    ]

    # Create a DataFrame with sample data and column headers
    data = {
        # Required fields with sample data
        'branch': ['Zirobwe'],
        'mem_reg_date': ['2025-04-10'],
        'member_number': ['ZIR20250001'],
        'first_name': ['John'],
        'surname': ['Doe'],
        'gender': ['Male'],
        'dob': ['1985-06-15'],
        'id_number': ['CM12345678PP9A'],
        'card_number': ['12345678'],
        'phone': ['256701234567'],
        'kin_sname': ['Smith'],
        'kin_fname': ['Jane'],
        'kin_phone': ['256712345678'],

        # Optional fields with examples
        'middle_name1': ['Robert'],
        'middle_name2': [''],
        'salutation': ['Mr'],
        'marital_status': ['Married'],
        'is_pwd': [False],
        'home_ownership': ['Owned'],
        'nin_village': ['Kampala'],
        'nin_parish': ['Central'],
        'nin_s_county': [''],
        'nin_county': [''],
        'nin_district': ['Kampala'],
        'phone_number2': [''],
        'email': ['john.doe@example.com'],
        'village': ['Kampala'],
        'parish': ['Central'],
        's_county': [''],
        'county': [''],
        'district': ['Kampala'],
        'sao_zone': ['Bamunanika'],
        'kin_sname2': [''],
        'kin_address': ['Kampala'],
        'kin_relationship': ['Spouse'],
        'employment': ['Self-employed'],
        'occupation': ['Farmer'],
        'employer_name': [''],
        'employer_address': [''],
        'employer_phone1': [''],
        'employer_phone2': [''],
        'income_frequency': ['Monthly'],
        'income_per_month': [300000]
    }

    # Create DataFrame
    df = pd.DataFrame(data)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')

        # Get the worksheet
        worksheet = writer.sheets['Template']

        # Style definitions
        required_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Light yellow
        required_font = Font(bold=True, color="000000")  # Bold black font

        # Adjust widths + highlight required headers
        for idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = 20

            cell = worksheet[f"{col_letter}1"]  # Header cell
            if col in required_fields:
                cell.fill = required_fill
                cell.font = required_font

    # Prepare response
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customer_import_template.xlsx'

    return response

# Import necessary libraries
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ValidationError
import pandas as pd
import os
import re
from datetime import datetime
from .models import Customer

import os, re
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from staff_management.models import Branch
from .models import Customer  # adjust if Customer is in another app

@login_required
def excel_import(request):
    """View to handle Excel file uploads for Customer data import"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Check file extension
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "Uploaded file must be an Excel file (.xls or .xlsx)")
            return redirect('excel_import')
            
        # Save the file temporarily
        fs = FileSystemStorage(location=os.path.join(os.getcwd(), 'temp'))
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        
        try:
            # Read Excel file with pandas
            df = pd.read_excel(file_path)
            
            records_imported = 0
            failed_records = []
            
            # Process each row
            for index, row in df.iterrows():
                try:
                    customer_data = {}
                    
                    # Dates
                    if 'mem_reg_date' in row and pd.notna(row['mem_reg_date']):
                        try:
                            customer_data['mem_reg_date'] = pd.to_datetime(row['mem_reg_date']).date()
                        except Exception:
                            raise ValidationError("Invalid registration date format")
                            
                    if 'dob' in row and pd.notna(row['dob']):
                        try:
                            customer_data['dob'] = pd.to_datetime(row['dob']).date()
                        except Exception:
                            raise ValidationError("Invalid date of birth format")
                    
                    # Boolean fields
                    if 'is_pwd' in row:
                        if isinstance(row['is_pwd'], bool):
                            customer_data['is_pwd'] = row['is_pwd']
                        elif isinstance(row['is_pwd'], str):
                            customer_data['is_pwd'] = row['is_pwd'].lower() in ['yes', 'true', '1', 'y']
                        elif isinstance(row['is_pwd'], (int, float)):
                            customer_data['is_pwd'] = bool(row['is_pwd'])
                    
                    # Text fields
                    text_fields = [
                        'member_number', 'email', 'salutation', 'first_name', 
                        'middle_name1', 'middle_name2', 'surname', 'gender', 'marital_status',
                        'home_ownership', 'id_number', 'card_number', 'nin_village', 'nin_parish',
                        'nin_s_county', 'nin_county', 'nin_district', 'phone', 
                        'phone_number2', 'village', 'parish', 's_county', 'county', 'district',
                        'sao_zone', 'kin_sname', 'kin_fname', 'kin_sname2', 'kin_phone',
                        'kin_address', 'kin_relationship', 'employment', 'occupation',
                        'employer_name', 'employer_address', 'employer_phone1', 
                        'employer_phone2', 'income_frequency'
                    ]
                    
                    for field in text_fields:
                        if field in row and pd.notna(row[field]):
                            customer_data[field] = str(row[field]).strip()
                    
                    # Numeric fields
                    if 'income_per_month' in row and pd.notna(row['income_per_month']):
                        try:
                            customer_data['income_per_month'] = float(row['income_per_month'])
                        except ValueError:
                            raise ValidationError("Invalid income value")
                    
                    # ✅ Branch lookup (ForeignKey instead of choices)
                    if 'branch' in row and pd.notna(row['branch']):
                        try:
                            # Adjust depending on Excel: is it branch name, code, or ID?
                            branch_obj = Branch.objects.get(name=row['branch'].strip())
                            customer_data['branch'] = branch_obj
                        except Branch.DoesNotExist:
                            raise ValidationError(f"Invalid branch: {row['branch']}")
                    
                    # Regex validations
                    if 'id_number' in customer_data:
                        if not re.match(r'^(CF|CM)[A-Z0-9]{11}[A-Z]$', customer_data['id_number']):
                            raise ValidationError("ID number format is invalid")
                    
                    if 'phone' in customer_data:
                        if not re.match(r'^256\d{9}$', customer_data['phone']):
                            raise ValidationError("Phone number format is invalid")
                    
                    if 'phone_number2' in customer_data and customer_data['phone_number2']:
                        if not re.match(r'^256\d{9}$', customer_data['phone_number2']):
                            raise ValidationError("Second phone number format is invalid")
                    
                    if 'kin_phone' in customer_data:
                        if not re.match(r'^256\d{9}$', customer_data['kin_phone']):
                            raise ValidationError("Kin phone number format is invalid")
                    
                    # Required fields
                    required_fields = [
                        'branch', 'mem_reg_date', 'member_number', 'first_name', 
                        'surname', 'gender', 'dob', 'id_number', 'card_number',
                        'phone', 'kin_sname', 'kin_fname', 'kin_phone'
                    ]
                    
                    for field in required_fields:
                        if field not in customer_data or not customer_data[field]:
                            raise ValidationError(f"Required field '{field}' is missing")
                    
                    # Save customer
                    customer = Customer(**customer_data)
                    customer.full_clean()
                    customer.save()
                    records_imported += 1
                    
                except Exception as e:
                    failed_records.append({
                        'row': index + 2,
                        'error': str(e)
                    })
            
            # Cleanup
            if os.path.exists(file_path):
                os.remove(file_path)
            
            # Messages
            if records_imported > 0:
                messages.success(request, f"Successfully imported {records_imported} customer(s)")
            
            if failed_records:
                failed_rows = ", ".join([str(record['row']) for record in failed_records[:5]])
                if len(failed_records) > 5:
                    failed_rows += f" and {len(failed_records) - 5} more"
                messages.warning(request, f"Failed to import {len(failed_records)} customer(s). Check rows: {failed_rows}")
                request.session['failed_records'] = failed_records
            
            return redirect('excel_import')
            
        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return redirect('excel_import')
    
    return render(request, 'import.html')
